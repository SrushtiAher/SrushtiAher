from flask import Flask, request, send_from_directory, jsonify
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
import os
import logging

# Initialize logging
logging.basicConfig(level=logging.DEBUG)

app = Flask(__name__)
CORS(app)

UPLOAD_FOLDER = 'C:\\Users\\Sagar\\Desktop\\New folder\\uploaded_files'
USER_DATA_PATH = "C:\\Users\\Sagar\\Desktop\\New folder\\user_data.xlsx"
LOG_FILE_PATH = "C:\\Users\\Sagar\\Desktop\\New folder\\activity_logs.txt"

# New Dictionary to keep track of user storage
user_storage = {}  # userID: {"limit": storage_limit, "used": used_storage}

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

@app.route('/log_activity', methods=['POST'])
def log_activity():
    log_data = request.json
    with open(LOG_FILE_PATH, "a") as log_file:
        log_file.write(f"{log_data['timestamp']} - {log_data['action']} - {log_data['fileName']} - {log_data['username']}\n")
    return jsonify({"message": "Log recorded"}), 200

@app.route('/upload', methods=['POST'])
def upload_file():
    user_id = "some_user_id"  # Replace this with actual user ID from your authentication system
    if user_id not in user_storage:
        user_storage[user_id] = {"limit": 500 * 1024 * 1024, "used": 0}  # Default to 500MB

    total_size = 0
    for filename in os.listdir(UPLOAD_FOLDER):
        file_path = os.path.join(UPLOAD_FOLDER, filename)
        total_size += os.path.getsize(file_path)

    incoming_file = request.files['file']
    incoming_file_size = len(incoming_file.read())
    incoming_file.seek(0)

    if total_size + incoming_file_size > user_storage[user_id]["limit"]:
        return 'Storage limit reached', 400

    user_storage[user_id]["used"] += incoming_file_size

    if user_storage[user_id]["used"] >= 0.9 * user_storage[user_id]["limit"]:
        print("90% storage limit reached")  # Replace this with actual notification logic

    file = incoming_file
    if file:
        file_path = os.path.join(UPLOAD_FOLDER, file.filename)
        if os.path.exists(file_path):
            return 'File already exists', 409
        else:
            file.save(file_path)
            return 'File uploaded successfully', 200
    else:
        return 'No file uploaded', 400

@app.route('/set_plan/<int:plan>', methods=['POST'])
def set_plan(plan):
    user_id = "some_user_id"  # Replace this with actual user ID from your authentication system
    if user_id not in user_storage:
        user_storage[user_id] = {"limit": 500 * 1024 * 1024, "used": 0}  # Default to 500MB
    
    if plan == 500:
        user_storage[user_id]["limit"] = 500 * 1024 * 1024  # 500MB
    elif plan == 1000:
        user_storage[user_id]["limit"] = 1000 * 1024 * 1024  # 1GB
    elif plan == 5000:
        user_storage[user_id]["limit"] = 5000 * 1024 * 1024  # 5GB
    else:
        return 'Invalid plan selected', 400
    
    return 'Plan updated successfully', 200

@app.route('/list', methods=['GET'])
def list_files():
    files = os.listdir(UPLOAD_FOLDER)
    file_data = []

    for file in files:
        file_path = os.path.join(UPLOAD_FOLDER, file)
        file_size = os.path.getsize(file_path)
        file_data.append({"name": file, "size": file_size})

    return jsonify(file_data)

@app.route('/download/<filename>', methods=['GET'])
def download_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename, as_attachment=True)

@app.route('/delete/<filename>', methods=['DELETE'])
def delete_file(filename):
    file_path = os.path.join(UPLOAD_FOLDER, filename)
    if os.path.exists(file_path):
        os.remove(file_path)
        return 'File deleted successfully', 200
    else:
        return 'File not found', 404

@app.route('/login', methods=['POST'])
def login():
    logging.debug("Received request: %s", request.form)
    user_id = request.form.get('userId')
    password = request.form.get('password')

    if not user_id or not password:
        return jsonify({"status": "failure", "message": "Missing userID or password"}), 400

    wb = load_workbook(USER_DATA_PATH)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2, values_only=True):
        stored_user_id, stored_password = row
        if stored_user_id == user_id and stored_password == password:
            return jsonify({"status": "success"}), 200

    return jsonify({"status": "failure", "message": "Invalid userID or password"}), 401

@app.route('/signup', methods=['POST'])
def signup():
    logging.debug("Received request: %s", request.form)
    user_id = request.form.get('userId')
    password = request.form.get('password')

    if not user_id or not password:
        return jsonify({"status": "failure", "message": "Missing userId or password"}), 400

    if os.path.exists(USER_DATA_PATH):
        wb = load_workbook(USER_DATA_PATH)
    else:
        wb = Workbook()
    ws = wb.active

    ws.append([user_id, password])
    wb.save(USER_DATA_PATH)

    return jsonify({"message": "User created"}), 201

@app.route('/logs', methods=['GET'])
def get_logs():
    logs = []
    try:
        with open(LOG_FILE_PATH, "r") as log_file:
            for line in log_file:
                parts = line.strip().split(" - ")
                if len(parts) == 4:
                    timestamp, action, filename, username = parts
                    logs.append({"timestamp": timestamp, "action": action, "filename": filename, "username": username})
    except FileNotFoundError:
        return jsonify({"error": "Log file not found"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    return jsonify(logs)


if __name__ == '__main__':
    app.run(port=8000)
